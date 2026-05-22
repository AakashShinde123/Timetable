"""Attendance: file-import + live eSSL/ZKTeco device sync (pyzk)."""
import csv
import io
import logging
from datetime import datetime, timezone, timedelta
from typing import Optional, List, Dict, Any
from fastapi import APIRouter, Depends, HTTPException, Body, UploadFile, File, Form

from db import db
from deps import get_current_user
from models import Attendance, serialize_doc

router = APIRouter(prefix="/api", tags=["attendance"])
log = logging.getLogger(__name__)


# ---- Resolving raw device user ids → our teacher ids -------------------------

async def _build_user_to_teacher_map(school_id: str) -> Dict[str, dict]:
    """Match eSSL device user-id (or name) to a teacher row.
    Priority: teacher.essl_user_id == raw_user_id → teacher.email → teacher.name (loose)."""
    teachers = await db.teachers.find({"school_id": school_id}, {"_id": 0}).to_list(1000)
    by_essl, by_name, by_email = {}, {}, {}
    for t in teachers:
        if t.get('essl_user_id'):
            by_essl[str(t['essl_user_id'])] = t
        if t.get('name'):
            by_name[t['name'].strip().lower()] = t
        if t.get('email'):
            by_email[t['email'].strip().lower()] = t
    return {"by_essl": by_essl, "by_name": by_name, "by_email": by_email, "all": teachers}


def _resolve_teacher(raw_uid: str, raw_name: str, maps: Dict[str, dict]) -> Optional[dict]:
    if raw_uid and str(raw_uid) in maps['by_essl']:
        return maps['by_essl'][str(raw_uid)]
    if raw_name:
        key = raw_name.strip().lower()
        if key in maps['by_email']:
            return maps['by_email'][key]
        if key in maps['by_name']:
            return maps['by_name'][key]
    return None


def _save_punches(school_id: str, source: str, device_id: Optional[str], rows: List[Dict[str, Any]]):
    """rows: [{raw_user_id, raw_user_name, date, time, punch_type, status}]"""
    async def go():
        maps = await _build_user_to_teacher_map(school_id)
        inserted, skipped, unmapped = 0, 0, 0
        for r in rows:
            if not r.get('date') or not r.get('time'):
                skipped += 1; continue
            teacher = _resolve_teacher(r.get('raw_user_id', ''), r.get('raw_user_name', ''), maps)
            # Dedupe: same school + teacher/raw_uid + date + time + punch_type
            dedupe_q = {
                "school_id": school_id,
                "date": r['date'],
                "time": r['time'],
                "punch_type": r.get('punch_type', 'in'),
                "raw_user_id": str(r.get('raw_user_id', '')),
            }
            if await db.attendance.find_one(dedupe_q):
                skipped += 1; continue
            doc = Attendance(
                school_id=school_id,
                teacher_id=(teacher or {}).get('id'),
                raw_user_id=str(r.get('raw_user_id', '')),
                raw_user_name=r.get('raw_user_name', ''),
                date=r['date'], time=r['time'],
                punch_type=r.get('punch_type', 'in'),
                source=source, device_id=device_id,
                status=r.get('status', ''),
            ).model_dump()
            ttl_at = doc['created_at']  # native datetime — used by the TTL index
            doc['created_at'] = doc['created_at'].isoformat()
            doc['ttl_at'] = ttl_at
            await db.attendance.insert_one(doc)
            inserted += 1
            if not teacher:
                unmapped += 1
        return {"inserted": inserted, "skipped": skipped, "unmapped": unmapped, "total_rows": len(rows)}
    return go


def _parse_punch_row(row: Dict[str, Any]) -> Dict[str, Any]:
    """Normalise one row (dict from csv.DictReader or xlsx) into our schema.
    Accepts common eSSL/ZK export shapes. Keys are case-insensitive."""
    lower = {(k or '').strip().lower(): (v if v is not None else '') for k, v in row.items()}
    raw_uid = lower.get('userid') or lower.get('user id') or lower.get('emp id') or lower.get('employee id') or lower.get('id') or ''
    raw_name = lower.get('name') or lower.get('employee name') or lower.get('user') or ''
    # Date / time can be combined ("DateTime") or separate
    date_val = lower.get('date') or lower.get('punchdate') or ''
    time_val = lower.get('time') or lower.get('punchtime') or ''
    dt_val = lower.get('datetime') or lower.get('date time') or lower.get('punch time') or ''
    date, time = '', ''
    if dt_val:
        for fmt in ('%Y-%m-%d %H:%M:%S', '%d/%m/%Y %H:%M:%S', '%m/%d/%Y %H:%M:%S',
                    '%Y-%m-%d %H:%M', '%d-%m-%Y %H:%M:%S'):
            try:
                dt = datetime.strptime(str(dt_val).strip(), fmt)
                date, time = dt.strftime('%Y-%m-%d'), dt.strftime('%H:%M:%S'); break
            except ValueError:
                continue
    if not date and date_val:
        for fmt in ('%Y-%m-%d', '%d/%m/%Y', '%m/%d/%Y', '%d-%m-%Y'):
            try:
                date = datetime.strptime(str(date_val).strip(), fmt).strftime('%Y-%m-%d'); break
            except ValueError:
                continue
    if not time and time_val:
        for fmt in ('%H:%M:%S', '%H:%M', '%I:%M %p', '%I:%M:%S %p'):
            try:
                time = datetime.strptime(str(time_val).strip(), fmt).strftime('%H:%M:%S'); break
            except ValueError:
                continue
    status = (lower.get('status') or '').strip().lower()
    punch = 'in' if status in ('in', 'check-in', 'checkin', '0') else 'out' if status in ('out', 'check-out', 'checkout', '1') else 'in'
    return {
        "raw_user_id": str(raw_uid).strip(),
        "raw_user_name": str(raw_name).strip(),
        "date": date,
        "time": time,
        "punch_type": punch,
        "status": status,
    }


# ---- Endpoints ---------------------------------------------------------------

@router.get("/schools/{school_id}/attendance")
async def list_attendance(school_id: str, date_from: Optional[str] = None,
                          date_to: Optional[str] = None, teacher_id: Optional[str] = None,
                          limit: int = 1000, offset: int = 0,
                          user: dict = Depends(get_current_user)):
    q: Dict[str, Any] = {"school_id": school_id}
    if date_from or date_to:
        rng = {}
        if date_from: rng["$gte"] = date_from
        if date_to: rng["$lte"] = date_to
        q["date"] = rng
    if teacher_id:
        q["teacher_id"] = teacher_id
    limit = max(1, min(limit, 5000))
    offset = max(0, offset)
    total = await db.attendance.count_documents(q)
    cursor = db.attendance.find(q, {"_id": 0, "ttl_at": 0}).sort([("date", -1), ("time", -1)]).skip(offset).limit(limit)
    docs = await cursor.to_list(limit)
    return {"total": total, "limit": limit, "offset": offset,
            "items": [serialize_doc(d) for d in docs]}


@router.delete("/schools/{school_id}/attendance/{att_id}")
async def delete_attendance(school_id: str, att_id: str, user: dict = Depends(get_current_user)):
    await db.attendance.delete_one({"id": att_id, "school_id": school_id})
    return {"ok": True}


@router.post("/schools/{school_id}/attendance/import-file")
async def import_attendance_file(school_id: str, file: UploadFile = File(...),
                                  device_id: Optional[str] = Form(None),
                                  user: dict = Depends(get_current_user)):
    """Parse a CSV (or .xlsx) export from eSSL software and persist punches."""
    contents = await file.read()
    fname = (file.filename or '').lower()
    rows: List[Dict[str, Any]] = []
    try:
        if fname.endswith('.xlsx') or fname.endswith('.xls'):
            import openpyxl
            wb = openpyxl.load_workbook(io.BytesIO(contents), data_only=True)
            ws = wb.active
            headers = [str(c.value or '').strip() for c in next(ws.iter_rows(min_row=1, max_row=1))]
            for r in ws.iter_rows(min_row=2, values_only=True):
                row_dict = {headers[i]: r[i] for i in range(min(len(headers), len(r)))}
                rows.append(_parse_punch_row(row_dict))
        else:
            text = contents.decode('utf-8', errors='replace')
            reader = csv.DictReader(io.StringIO(text))
            for row in reader:
                rows.append(_parse_punch_row(row))
    except Exception as e:
        raise HTTPException(400, f"Could not parse file: {e}")
    if not rows:
        raise HTTPException(400, "No rows parsed. Check column headers: UserID/Name/Date/Time or DateTime")
    result = await _save_punches(school_id, "file-upload", device_id, rows)()
    result['source_file'] = file.filename
    return result


@router.post("/schools/{school_id}/attendance/sync-essl")
async def sync_essl(school_id: str, payload: Dict[str, Any] = Body(default={}),
                    user: dict = Depends(get_current_user)):
    """Live pull from an eSSL/ZKTeco device. Body: {device_id, days_back?:int} or {ip,port,password}.
    If device unreachable, returns 502 so the UI can prompt the admin to fall back to file upload."""
    days_back = int(payload.get('days_back', 7))
    device_id = payload.get('device_id')
    if device_id:
        dev = await db.essl_devices.find_one({"id": device_id, "school_id": school_id}, {"_id": 0})
        if not dev:
            raise HTTPException(404, "eSSL device not found")
    else:
        if not payload.get('ip'):
            raise HTTPException(400, "device_id or ip required")
        dev = {
            "id": None, "school_id": school_id, "name": payload.get('name', 'adhoc'),
            "ip": payload['ip'], "port": int(payload.get('port', 4370)),
            "password": int(payload.get('password', 0)),
            "timeout": int(payload.get('timeout', 8)),
            "force_udp": bool(payload.get('force_udp', False)),
            "ommit_ping": bool(payload.get('ommit_ping', True)),
        }

    try:
        from zk import ZK
    except ImportError:
        raise HTTPException(500, "pyzk not installed on server")

    zk = ZK(dev['ip'], port=dev['port'], timeout=dev['timeout'],
            password=dev['password'], force_udp=dev['force_udp'],
            ommit_ping=dev['ommit_ping'])
    conn = None
    try:
        conn = zk.connect()
        users = {str(u.user_id): u.name for u in (conn.get_users() or [])}
        attendances = conn.get_attendance() or []
    except Exception as e:
        log.warning("eSSL connect failed for school %s: %s", school_id, e)
        raise HTTPException(502, f"Could not reach eSSL device at {dev['ip']}:{dev['port']} — {e}")
    finally:
        try:
            if conn:
                conn.disconnect()
        except Exception:
            pass

    cutoff = datetime.now(timezone.utc) - timedelta(days=days_back)
    rows: List[Dict[str, Any]] = []
    for a in attendances:
        ts = a.timestamp
        if ts and ts.tzinfo is None:
            ts = ts.replace(tzinfo=timezone.utc)
        if ts and ts < cutoff:
            continue
        rows.append({
            "raw_user_id": str(a.user_id),
            "raw_user_name": users.get(str(a.user_id), ''),
            "date": ts.strftime('%Y-%m-%d') if ts else '',
            "time": ts.strftime('%H:%M:%S') if ts else '',
            "punch_type": 'in' if int(getattr(a, 'punch', 0) or 0) % 2 == 0 else 'out',
            "status": str(getattr(a, 'status', '')),
        })
    result = await _save_punches(school_id, "essl-network", dev.get('id'), rows)()
    result['device'] = {"ip": dev['ip'], "name": dev.get('name')}
    result['users_on_device'] = len(users)
    return result


@router.post("/schools/{school_id}/attendance/manual")
async def add_manual_attendance(school_id: str, payload: Dict[str, Any] = Body(...),
                                 user: dict = Depends(get_current_user)):
    """Admin manual punch entry. Body: {teacher_id, date, time, punch_type}."""
    if not payload.get('teacher_id') or not payload.get('date') or not payload.get('time'):
        raise HTTPException(400, "teacher_id, date, time required")
    doc = Attendance(
        school_id=school_id, teacher_id=payload['teacher_id'],
        raw_user_id='', raw_user_name=payload.get('raw_user_name', ''),
        date=payload['date'], time=payload['time'],
        punch_type=payload.get('punch_type', 'in'),
        source='manual',
    ).model_dump()
    ttl_at = doc['created_at']
    doc['created_at'] = doc['created_at'].isoformat()
    doc['ttl_at'] = ttl_at
    await db.attendance.insert_one(doc)
    return serialize_doc(doc)


@router.get("/schools/{school_id}/attendance/summary")
async def attendance_summary(school_id: str, date: Optional[str] = None,
                              user: dict = Depends(get_current_user)):
    """Per-teacher summary for a given date (default = today)."""
    if not date:
        date = datetime.now(timezone.utc).strftime('%Y-%m-%d')
    teachers = await db.teachers.find({"school_id": school_id}, {"_id": 0}).to_list(1000)
    punches = await db.attendance.find({"school_id": school_id, "date": date}, {"_id": 0}).to_list(5000)
    by_teacher: Dict[str, List[dict]] = {}
    unmapped = []
    for p in punches:
        if p.get('teacher_id'):
            by_teacher.setdefault(p['teacher_id'], []).append(p)
        else:
            unmapped.append(p)
    summary = []
    for t in teachers:
        ps = sorted(by_teacher.get(t['id'], []), key=lambda x: x['time'])
        if ps:
            in_time = ps[0]['time']
            out_time = ps[-1]['time'] if len(ps) > 1 else None
        else:
            in_time = out_time = None
        summary.append({
            "teacher_id": t['id'], "teacher_name": t.get('name'),
            "abbreviation": t.get('abbreviation'),
            "punches": len(ps),
            "first_in": in_time, "last_out": out_time,
            "present": bool(ps),
        })
    summary.sort(key=lambda x: (not x['present'], x['teacher_name'] or ''))
    return {
        "date": date,
        "total_teachers": len(teachers),
        "present": sum(1 for s in summary if s['present']),
        "absent": sum(1 for s in summary if not s['present']),
        "unmapped_punches": len(unmapped),
        "rows": summary,
    }
